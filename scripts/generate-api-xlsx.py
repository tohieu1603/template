#!/usr/bin/env python3
"""
Generate API reference XLSX for each template.
Each xlsx has sheets: Routes, RBAC Matrix, Seed Data
"""
import os
import re
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

TEMPLATES = [
    ("01-ecommerce", 3000),
    ("02-blog", 3001),
    ("03-portfolio", 3001),
    ("05-saas", 3002),
    ("06-booking", 3003),
    ("07-restaurant", 3004),
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PUBLIC_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
AUTH_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
RBAC_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def parse_routes(project_dir):
    """Parse all *.routes.ts files and extract route definitions."""
    routes = []
    route_files = sorted(glob.glob(os.path.join(project_dir, "src/modules/*/*.routes.ts")))

    for rf in route_files:
        module = os.path.basename(os.path.dirname(rf))
        with open(rf, "r") as f:
            content = f.read()

        # Find route registrations
        pattern = r"router\.(get|post|put|patch|delete)\(\s*['\"]([^'\"]+)['\"]"
        matches = re.finditer(pattern, content, re.IGNORECASE)

        for m in matches:
            method = m.group(1).upper()
            path = m.group(2)

            # Get the full line for middleware analysis
            line_start = content.rfind('\n', 0, m.start()) + 1
            line_end = content.find('\n', m.end())
            full_line = content[line_start:line_end] if line_end > 0 else content[line_start:]

            # Determine auth level
            has_auth = 'auth()' in full_line
            rbac_match = re.search(r"rbac\(['\"]([^'\"]+)['\"]\)", full_line)
            rbac_perm = rbac_match.group(1) if rbac_match else ""

            if rbac_perm:
                auth_level = f"RBAC: {rbac_perm}"
            elif has_auth:
                auth_level = "AUTH"
            else:
                auth_level = "Public"

            # Extract DTO class name
            dto_match = re.search(r"validateDto\((\w+)", full_line)
            dto_name = dto_match.group(1) if dto_match else ""

            # Handler function name
            handler_match = re.search(r",\s*(\w+)\s*\)", full_line)
            handler = handler_match.group(1) if handler_match else ""

            routes.append({
                "module": module,
                "method": method,
                "path": path,
                "auth": auth_level,
                "permission": rbac_perm,
                "dto": dto_name,
                "handler": handler,
            })

    return routes


def parse_dto_fields(project_dir, dto_name):
    """Find DTO class and extract field names."""
    dto_files = glob.glob(os.path.join(project_dir, "src/modules/*/dto/*.dto.ts"))
    for df in dto_files:
        with open(df, "r") as f:
            content = f.read()
        # Find the class
        class_match = re.search(rf"export class {re.escape(dto_name)}\s*(?:extends\s+\w+\s*)?\{{(.*?)\}}", content, re.DOTALL)
        if class_match:
            body = class_match.group(1)
            # Extract field names (lines with : type)
            fields = re.findall(r"(\w+)\s*[\?]?\s*:", body)
            # Filter out decorator names
            fields = [f for f in fields if not f[0].isupper() and f not in ('type',)]
            return fields
    return []


def get_route_prefix(app_ts_path, module_name):
    """Get the URL prefix for a module from app.ts."""
    if not os.path.exists(app_ts_path):
        return f"/api/v1/{module_name}s"
    with open(app_ts_path, "r") as f:
        content = f.read()
    # Match: app.use(`${prefix}/xxx`, yyyRoutes)
    # Find the import name for this module
    patterns = [
        rf"app\.use\(`\${{prefix}}/([^`]+)`\s*,\s*\w*{re.escape(module_name.replace('-', ''))}",
        rf"app\.use\(`\${{prefix}}/([^`]+)`\s*,\s*\w*{re.escape(module_name)}",
    ]
    for p in patterns:
        m = re.search(p, content, re.IGNORECASE)
        if m:
            return f"/api/v1/{m.group(1)}"
    return f"/api/v1/{module_name}"


def create_xlsx(template_name, port):
    project_dir = os.path.join(BASE, template_name)
    if not os.path.isdir(os.path.join(project_dir, "src")):
        print(f"  SKIP {template_name} (no src/)")
        return

    app_ts = os.path.join(project_dir, "src/app.ts")
    routes = parse_routes(project_dir)

    if not routes:
        print(f"  SKIP {template_name} (no routes found)")
        return

    wb = Workbook()

    # ============ Sheet 1: All Routes ============
    ws = wb.active
    ws.title = "API Routes"

    headers = ["#", "Module", "Method", "Full Path", "Auth Level", "Permission", "DTO / Params", "Handler"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    row = 2
    for i, r in enumerate(routes, 1):
        prefix = get_route_prefix(app_ts, r["module"])
        full_path = f"{prefix}{r['path']}" if r['path'] != '/' else prefix

        # Get DTO fields as param hint
        params = ""
        if r["dto"]:
            fields = parse_dto_fields(project_dir, r["dto"])
            if fields:
                params = f"{r['dto']}: {{ {', '.join(fields)} }}"
            else:
                params = r["dto"]

        values = [i, r["module"], r["method"], full_path, r["auth"], r["permission"], params, r["handler"]]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Color by auth level
            if col == 5:
                if "Public" in str(v):
                    cell.fill = PUBLIC_FILL
                elif "AUTH" == str(v):
                    cell.fill = AUTH_FILL
                elif "RBAC" in str(v):
                    cell.fill = RBAC_FILL

            # Method color
            if col == 3:
                colors = {"GET": "2E75B6", "POST": "548235", "PUT": "BF8F00", "PATCH": "BF8F00", "DELETE": "C00000"}
                cell.font = Font(bold=True, color=colors.get(str(v), "000000"))

        row += 1

    # Column widths
    widths = [5, 18, 8, 50, 25, 30, 50, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Freeze header
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{row - 1}"

    # ============ Sheet 2: RBAC Matrix ============
    ws2 = wb.create_sheet("RBAC Matrix")

    # Collect unique modules and permissions
    modules = sorted(set(r["module"] for r in routes))
    actions = ["view", "create", "update", "delete"]

    headers2 = ["Module"] + actions
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h.upper())
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    for i, mod in enumerate(modules, 2):
        ws2.cell(row=i, column=1, value=mod).border = THIN_BORDER
        for j, action in enumerate(actions, 2):
            perm = f"{mod.replace('-', '_')}.{action}"
            # Check if any route uses this permission
            has = any(r["permission"] == perm for r in routes)
            # Also check without module transform
            if not has:
                has = any(perm in r["permission"] for r in routes if r["permission"])
            cell = ws2.cell(row=i, column=j, value="✓" if has else "")
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
            if has:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws2.column_dimensions['A'].width = 20
    for c in 'BCDE':
        ws2.column_dimensions[c].width = 12
    ws2.freeze_panes = "B2"

    # ============ Sheet 3: Summary ============
    ws3 = wb.create_sheet("Summary")

    summary_data = [
        ["Template", template_name],
        ["Port", port],
        ["Base URL", f"http://localhost:{port}/api/v1"],
        ["Swagger", f"http://localhost:{port}/api/v1/docs"],
        ["Total Modules", len(modules)],
        ["Total Routes", len(routes)],
        ["Public Routes", sum(1 for r in routes if r["auth"] == "Public")],
        ["Auth Routes", sum(1 for r in routes if r["auth"] == "AUTH")],
        ["RBAC Routes", sum(1 for r in routes if "RBAC" in r["auth"])],
        ["GET", sum(1 for r in routes if r["method"] == "GET")],
        ["POST", sum(1 for r in routes if r["method"] == "POST")],
        ["PUT", sum(1 for r in routes if r["method"] == "PUT")],
        ["PATCH", sum(1 for r in routes if r["method"] == "PATCH")],
        ["DELETE", sum(1 for r in routes if r["method"] == "DELETE")],
    ]

    for i, (key, val) in enumerate(summary_data, 1):
        cell_k = ws3.cell(row=i, column=1, value=key)
        cell_k.font = Font(bold=True)
        cell_k.border = THIN_BORDER
        cell_v = ws3.cell(row=i, column=2, value=val)
        cell_v.border = THIN_BORDER

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 40

    # Save
    output_dir = os.path.join(BASE, "plans")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"api-reference-{template_name}.xlsx")
    wb.save(output_path)
    print(f"  {template_name}: {len(routes)} routes → {output_path}")


if __name__ == "__main__":
    print("Generating API reference XLSX files...")
    for name, port in TEMPLATES:
        create_xlsx(name, port)
    print("Done!")
